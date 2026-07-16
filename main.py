from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Any
import psycopg2, psycopg2.extras, os, hashlib, secrets, time, uuid, io
from datetime import datetime
from dotenv import load_dotenv
import httpx
import xlrd, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

app = FastAPI(title="Bata Asset Audit API", version="1.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://batathai.github.io","http://localhost","http://127.0.0.1","https://fixasset.batathai.com"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    conn = psycopg2.connect(os.getenv("DATABASE_URL"), cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        yield conn
    finally:
        conn.close()

sessions = {}

def get_current_user(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.split(" ")[1]
    if token not in sessions:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return sessions[token]

# ════════════════════════════════════════════════════════════════
# MODELS
# ════════════════════════════════════════════════════════════════
class LoginRequest(BaseModel):
    employee_id: str
    password: str
    branch_id: str

class ScanLogCreate(BaseModel):
    session_id: int
    qr_key: str
    serial_found: Optional[str] = None
    serial_match: Optional[bool] = None
    condition: str = "good"
    remark: Optional[str] = None
    photo_url: Optional[str] = None

class UnmatchedCreate(BaseModel):
    session_id: int
    scanned_qr: str
    serial_no: Optional[str] = None
    name_guess: Optional[str] = None
    photo_url: Optional[str] = None
    remark: Optional[str] = None

class SessionCreate(BaseModel):
    branch_id: str
    audit_date: str
    name: Optional[str] = None

class ScanLogUpdate(BaseModel):
    condition: Optional[str] = None
    serial_found: Optional[str] = None
    hq_note: Optional[str] = None

class ScanLogPatchByQR(BaseModel):
    photo_url: Optional[str] = None
    remark: Optional[str] = None
    serial_found: Optional[str] = None
    serial_match: Optional[bool] = None
    condition: Optional[str] = None

class UnmatchedUpdate(BaseModel):
    status: str
    hq_note: Optional[str] = None
    matched_asset_code: Optional[str] = None
    serial_no: Optional[str] = None

class MergeSessionsRequest(BaseModel):
    primary_session_id: int
    merge_session_ids: List[int]
    branch_id: str

class ExportExcelRequest(BaseModel):
    title: str                     # เช่น "Overview Report"
    branch_label: str              # เช่น "Homepro Suvarnabhumi (0051435)" หรือ "12 สาขา"
    headers: List[str]
    rows: List[List[Any]]
    filename: str                  # เช่น "overview-0051435-20260716.xlsx"

# ════════════════════════════════════════════════════════════════
# CLOUDINARY HELPERS
# ════════════════════════════════════════════════════════════════
def _cloudinary_public_id(url: str) -> str:
    url = url.split("?")[0]
    if "/upload/" not in url:
        return ""
    after_upload = url.split("/upload/", 1)[1]
    parts = after_upload.split("/")
    if parts[0].startswith("v") and parts[0][1:].isdigit():
        parts = parts[1:]
    path = "/".join(parts)
    if "." in path.rsplit("/", 1)[-1]:
        path = path.rsplit(".", 1)[0]
    return path

def _cloudinary_sign(params: dict, api_secret: str) -> str:
    sorted_params = "&".join(
        f"{k}={v}" for k, v in sorted(params.items())
        if k not in ("file", "api_key", "resource_type")
    )
    return hashlib.sha1(f"{sorted_params}{api_secret}".encode()).hexdigest()

# ════════════════════════════════════════════════════════════════
# SCAN DELETE AUDIT LOG
# ต้องรัน migration_scan_delete_log.sql ใน Neon ก่อน endpoint กลุ่มนี้ถึงจะใช้ได้
# (ตาราง scan_delete_logs ไม่ได้ auto-create ตอน startup ตาม convention เดิมของไฟล์นี้
#  ที่ apply schema change ผ่าน Neon SQL editor เท่านั้น ดู docs/database.md)
# ════════════════════════════════════════════════════════════════
def _record_scan_delete(cur, scan_row, session_id, branch_id, user, source):
    """บันทึกทุกครั้งที่มีการลบ scan_log ไว้ใน scan_delete_logs ก่อน DELETE จริงเสมอ
    scan_row ต้องมี: id, asset_id, asset_code (or None), name (or None), qr_key (or None)
    source: 'hq' | 'branch' — ใช้แยกว่าลบจากฝั่งไหน"""
    cur.execute("""
        INSERT INTO scan_delete_logs
            (scan_log_id, session_id, branch_id, asset_code, asset_name, qr_key,
             deleted_by, deleted_by_role, source)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        scan_row.get("id"), session_id, branch_id,
        scan_row.get("asset_code"), scan_row.get("name"), scan_row.get("qr_key"),
        user.get("employee_id") or user.get("full_name") or str(user.get("user_id")),
        user.get("role"), source
    ))

# ════════════════════════════════════════════════════════════════
# AUTH
# ════════════════════════════════════════════════════════════════
@app.post("/auth/login")
def login(req: LoginRequest, db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("SELECT id, full_name, role, password_hash FROM users WHERE email = %s AND is_active = true", (req.employee_id,))
    user = cur.fetchone()
    pw_hash = hashlib.sha256(req.password.encode()).hexdigest()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not user["password_hash"]:
        cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (pw_hash, user["id"]))
        db.commit()
    elif user["password_hash"] != pw_hash:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = secrets.token_hex(32)
    sessions[token] = {"user_id": user["id"], "employee_id": req.employee_id, "branch_id": req.branch_id, "full_name": user["full_name"], "role": user["role"]}
    return {"token": token, "employee_id": req.employee_id, "full_name": user["full_name"], "branch_id": req.branch_id, "role": user["role"]}

@app.post("/auth/logout")
def logout(authorization: str = Header(None)):
    if authorization and authorization.startswith("Bearer "):
        sessions.pop(authorization.split(" ")[1], None)
    return {"ok": True}

# ════════════════════════════════════════════════════════════════
# ASSETS
# ════════════════════════════════════════════════════════════════
@app.get("/assets/lookup/{qr_key}")
def lookup_asset(qr_key: str, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("SELECT * FROM assets WHERE qr_key = %s", (qr_key,))
    asset = cur.fetchone()
    if not asset:
        return {"found": False, "qr_key": qr_key}
    return {"found": True, "asset": dict(asset)}

@app.get("/assets/branch/{branch_id}")
def assets_by_branch(branch_id: str, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    # ดึง master list ทั้งหมดของ branch (ใช้ location_code)
    cur.execute("""
        SELECT id, qr_key, asset_code, seq, name, serial_no,
               purchase_date, status, location_code
        FROM assets
        WHERE location_code = %s AND status = 'active' AND COALESCE(is_active, true) = true
        ORDER BY asset_code, seq
    """, (branch_id,))
    rows = cur.fetchall()

    # Fallback: ถ้าได้น้อยกว่า 10 ให้ดึงจาก scan_logs ที่ branch นี้เคย scan
    if len(rows) < 10:
        cur.execute("""
            SELECT DISTINCT a.id, a.qr_key, a.asset_code, a.seq, a.name,
                   a.serial_no, a.purchase_date, a.status, a.location_code
            FROM assets a
            JOIN scan_logs sl ON sl.asset_id = a.id
            JOIN audit_sessions s ON s.id = sl.session_id
            WHERE s.branch_id = %s AND COALESCE(a.is_active, true) = true
            UNION
            SELECT id, qr_key, asset_code, seq, name, serial_no,
                   purchase_date, status, location_code
            FROM assets
            WHERE location_code = %s AND COALESCE(is_active, true) = true
            ORDER BY asset_code, seq
        """, (branch_id, branch_id))
        rows = cur.fetchall()

    return {"assets": [dict(r) for r in rows]}

# ════════════════════════════════════════════════════════════════
# AUDIT SESSIONS
# audit_sessions columns: id, name, branch_id, audit_date,
#                         started_by, status, created_at, closed_at
# ════════════════════════════════════════════════════════════════
@app.post("/sessions")
def create_session(req: SessionCreate, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    name = req.name or f"Audit {req.branch_id} {req.audit_date}"
    # FIX (race condition): เดิมเช็คด้วย SELECT ก่อนแล้วค่อย INSERT แยกกัน 2 query
    # ถ้า Audit A กับ B login พร้อมกันเป๊ะๆ ทั้งคู่ SELECT ไม่เจอ session เดิมพร้อมกันได้
    # แล้ว INSERT ซ้ำกันทั้งคู่ -> ได้ session คนละอัน ทั้งที่ควรเป็น session เดียวกัน
    # แก้โดยใช้ INSERT...ON CONFLICT กับ partial unique index
    # idx_unique_open_session_per_branch_day (ดู migration_session_uniqueness.sql)
    # ซึ่ง Postgres รับประกัน atomicity ระดับ DB ให้เอง ไม่ต้องพึ่ง check-then-act ฝั่งแอปอีกต่อไป
    cur.execute("""
        INSERT INTO audit_sessions (name, branch_id, audit_date, started_by, status)
        VALUES (%s,%s,%s,%s,'open')
        ON CONFLICT (branch_id, audit_date) WHERE status != 'done' DO NOTHING
        RETURNING id
    """, (name, req.branch_id, req.audit_date, user["user_id"]))
    row = cur.fetchone()
    if row:
        db.commit()
        return {"session_id": row["id"], "reused": False}
    # เกิด conflict แปลว่ามี session เดิมอยู่แล้ว (อาจเป็นเพราะอีกคน insert ไปก่อนเสี้ยววินาที) -> ดึงมาใช้ร่วมกัน
    cur.execute("SELECT id FROM audit_sessions WHERE branch_id = %s AND audit_date = %s AND status != 'done'", (req.branch_id, req.audit_date))
    existing = cur.fetchone()
    db.commit()
    return {"session_id": existing["id"], "reused": True}

@app.get("/sessions/{session_id}/progress")
def session_progress(session_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("""
        SELECT
            s.id AS session_id, s.branch_id, s.name, s.status, s.audit_date,
            s.created_at, s.closed_at,
            COUNT(sl.id) AS scanned_count,
            (SELECT COUNT(*) FROM assets a
             WHERE a.location_code = s.branch_id AND a.status = 'active') AS total_assets,
            (SELECT COUNT(*) FROM unmatched_assets ua WHERE ua.session_id = s.id AND ua.status = 'pending') AS unmatched_count
        FROM audit_sessions s
        LEFT JOIN scan_logs sl ON sl.session_id = s.id
        WHERE s.id = %s
        GROUP BY s.id, s.branch_id, s.name, s.status, s.audit_date, s.created_at, s.closed_at
    """, (session_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Session not found")
    return dict(row)

@app.patch("/sessions/{session_id}/close")
def close_session(session_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("UPDATE audit_sessions SET status='done', closed_at=now() WHERE id=%s RETURNING id", (session_id,))
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="Session not found")
    db.commit()
    return {"ok": True}

# ════════════════════════════════════════════════════════════════
# SCAN LOGS
# scan_logs columns: id, session_id, asset_id, scanned_by,
#   scanned_at, serial_verified, serial_match, serial_found,
#   condition, remark, photo_url, hq_note, updated_at
# ════════════════════════════════════════════════════════════════
@app.post("/scans")
def create_scan(req: ScanLogCreate, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("SELECT id FROM assets WHERE qr_key = %s", (req.qr_key,))
    asset = cur.fetchone()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    # เช็คซ้ำข้าม session ในวันเดียวกัน branch เดียวกัน
    # ดึง scanned_by (ชื่อคนสแกน) + scanned_at ออกมาด้วย เพื่อส่งกลับให้ client
    # ตาม contract ใน API.md ({"error":"already_scanned","scanned_by":...,"scanned_at":...})
    # เดิมส่งแค่ detail เป็น string ธรรมดา ทำให้ฝั่ง index.html เช็ค data.error ไม่เจอ
    # และไม่ rollback การ์ดที่ขึ้น "Verified" แบบ optimistic ไปก่อนหน้า กลายเป็นบั๊ก
    # "สแกนซ้ำแล้วยังขึ้น Verified" ที่หน้างานเจอ
    cur.execute("""
        SELECT sl.id, sl.scanned_at, u.full_name AS scanned_by_name
        FROM scan_logs sl
        JOIN audit_sessions s ON s.id = sl.session_id
        LEFT JOIN users u ON u.id = sl.scanned_by
        WHERE sl.asset_id = %s
          AND s.branch_id = (SELECT branch_id FROM audit_sessions WHERE id = %s)
          AND s.audit_date = (SELECT audit_date FROM audit_sessions WHERE id = %s)
    """, (asset["id"], req.session_id, req.session_id))
    dup = cur.fetchone()
    if dup:
        raise HTTPException(status_code=409, detail={
            "error": "already_scanned",
            "scanned_by": dup["scanned_by_name"] or "unknown",
            "scanned_at": str(dup["scanned_at"])
        })
    # เช็คซ้ำเฉพาะ session นี้ (กรณี edge case ที่ query ข้างบนพลาด เช่น audit_date ไม่ตรงกันแบบไม่คาดคิด)
    cur.execute("""
        SELECT sl.id, sl.scanned_at, u.full_name AS scanned_by_name
        FROM scan_logs sl
        LEFT JOIN users u ON u.id = sl.scanned_by
        WHERE sl.session_id = %s AND sl.asset_id = %s
    """, (req.session_id, asset["id"]))
    dup2 = cur.fetchone()
    if dup2:
        raise HTTPException(status_code=409, detail={
            "error": "already_scanned",
            "scanned_by": dup2["scanned_by_name"] or "unknown",
            "scanned_at": str(dup2["scanned_at"])
        })
    cur.execute("""
        INSERT INTO scan_logs
            (session_id, asset_id, scanned_by, serial_found, serial_match, condition, remark, photo_url)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (req.session_id, asset["id"], user["user_id"], req.serial_found,
          req.serial_match, req.condition, req.remark, req.photo_url))
    scan_id = cur.fetchone()["id"]
    cur.execute("UPDATE audit_sessions SET status = 'on_process' WHERE id = %s AND status = 'open'", (req.session_id,))
    db.commit()
    return {"scan_id": scan_id, "asset_id": asset["id"]}

@app.get("/sessions/{session_id}/scans")
def get_scans(session_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("""
        SELECT sl.id, sl.scanned_at, sl.serial_match, sl.condition, sl.remark, sl.photo_url,
               a.qr_key, a.name, a.serial_no
        FROM scan_logs sl
        JOIN assets a ON a.id = sl.asset_id
        WHERE sl.session_id = %s
        ORDER BY sl.scanned_at DESC
    """, (session_id,))
    return {"scans": [dict(r) for r in cur.fetchall()]}

# ════════════════════════════════════════════════════════════════
# UNMATCHED ASSETS
# unmatched_assets columns: id, session_id, scanned_qr, serial_no,
#   name_guess, photo_url, scanned_by, branch_id, scanned_at,
#   status, hq_note, reviewed_at, matched_asset_code,
#   reviewed_by, updated_at
# ════════════════════════════════════════════════════════════════
@app.post("/unmatched")
def create_unmatched(req: UnmatchedCreate, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("""
        INSERT INTO unmatched_assets
            (session_id, scanned_qr, serial_no, name_guess, photo_url, scanned_by, branch_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (req.session_id, req.scanned_qr, req.serial_no, req.name_guess,
          req.photo_url, user["user_id"], user["branch_id"]))
    uid = cur.fetchone()["id"]
    db.commit()
    return {"unmatched_id": uid}

@app.get("/unmatched/pending")
def pending_unmatched(db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("""
        SELECT ua.*, u.email as auditor FROM unmatched_assets ua
        LEFT JOIN users u ON u.id = ua.scanned_by
        WHERE ua.status = 'pending' ORDER BY ua.scanned_at DESC
    """)
    return {"items": [dict(r) for r in cur.fetchall()]}

# ════════════════════════════════════════════════════════════════
# DASHBOARD — คำนวณจาก scan_logs จริง ไม่ใช้ v_session_progress
# ════════════════════════════════════════════════════════════════
@app.get("/dashboard/summary")
def dashboard_summary(db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("""
        SELECT
            s.id          AS session_id,
            s.branch_id,
            s.name,
            s.status,
            s.audit_date,
            s.created_at,
            s.closed_at,
            s.started_by,
            COUNT(sl.id)  AS scanned_count,
            (SELECT COUNT(*) FROM assets a
             WHERE a.location_code = s.branch_id
               AND a.status = 'active') AS total_assets,
            (SELECT COUNT(*) FROM unmatched_assets ua
             WHERE ua.session_id = s.id
               AND ua.status = 'pending') AS unmatched_count
        FROM audit_sessions s
        LEFT JOIN scan_logs sl ON sl.session_id = s.id
        GROUP BY s.id, s.branch_id, s.name, s.status,
                 s.audit_date, s.created_at, s.closed_at, s.started_by
        ORDER BY s.audit_date DESC
    """)
    sessions_data = [dict(r) for r in cur.fetchall()]

    # Fallback: ถ้า total_assets = 0 ให้ใช้จำนวน distinct asset ที่ scan ไปแล้ว
    for s in sessions_data:
        if not s["total_assets"]:
            cur.execute("""
                SELECT COUNT(DISTINCT asset_id) AS cnt FROM scan_logs WHERE session_id = %s
            """, (s["session_id"],))
            s["total_assets"] = cur.fetchone()["cnt"] or 0

    cur.execute("SELECT COUNT(*) as total FROM assets WHERE status='active'")
    total_assets = cur.fetchone()["total"]
    cur.execute("SELECT COUNT(*) as total FROM unmatched_assets WHERE status='pending'")
    pending_unmatched = cur.fetchone()["total"]

    return {
        "total_assets": total_assets,
        "pending_unmatched": pending_unmatched,
        "sessions": sessions_data
    }

# ════════════════════════════════════════════════════════════════
# HQ ENDPOINTS
# ════════════════════════════════════════════════════════════════
@app.get("/hq/scans")
def hq_get_scans(db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("""
        SELECT sl.id, sl.condition, sl.serial_found, sl.serial_match, sl.photo_url,
               sl.scanned_at, sl.hq_note, a.asset_code, a.seq, a.qr_key,
               a.name AS asset_name, a.serial_no AS serial_master,
               u.email AS auditor,
               s.id AS session_id, b.id AS branch_id, b.name AS branch_name
        FROM scan_logs sl
        LEFT JOIN assets a ON a.id = sl.asset_id
        LEFT JOIN users u ON u.id = sl.scanned_by
        LEFT JOIN audit_sessions s ON s.id = sl.session_id
        LEFT JOIN branches b ON b.id = s.branch_id
        ORDER BY sl.scanned_at DESC
    """)
    return {"scans": [dict(r) for r in cur.fetchall()]}

@app.patch("/hq/scans/{scan_id}")
def hq_update_scan(scan_id: int, req: ScanLogUpdate, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    fields, values = [], []
    if req.condition is not None:
        fields.append("condition = %s"); values.append(req.condition)
    if req.serial_found is not None:
        fields.append("serial_found = %s"); values.append(req.serial_found or None)
    if req.hq_note is not None:
        fields.append("hq_note = %s"); values.append(req.hq_note or None)
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    fields.append("updated_at = NOW()")
    values.append(scan_id)
    cur.execute(f"UPDATE scan_logs SET {', '.join(fields)} WHERE id = %s RETURNING id", values)
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="Scan log not found")
    db.commit()
    return {"ok": True, "scan_id": scan_id}

@app.delete("/hq/scans/{scan_id}")
def hq_delete_scan(scan_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    # ต้อง SELECT ข้อมูลก่อนลบเสมอ เพื่อเอาไปบันทึกใน scan_delete_logs
    # (ลบจริงแบบถาวร แต่เก็บ log แยกไว้ต่างหากตามที่ HQ ต้องการ audit trail)
    cur.execute("""
        SELECT sl.id, sl.session_id, a.id AS asset_id, a.asset_code, a.name, a.qr_key,
               s.branch_id
        FROM scan_logs sl
        JOIN assets a ON a.id = sl.asset_id
        LEFT JOIN audit_sessions s ON s.id = sl.session_id
        WHERE sl.id = %s
    """, (scan_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Scan not found")
    _record_scan_delete(cur, dict(row), row["session_id"], row["branch_id"], user, source="hq")
    cur.execute("DELETE FROM scan_logs WHERE id = %s", (scan_id,))
    db.commit()
    return {"ok": True}

# ════════════════════════════════════════════════════════════════
# BRANCH SELF-SERVICE SCAN DELETE (แก้บั๊ก "ลบในแอพสแกนแล้วข้อมูลไม่หาย")
# ต่างจาก /hq/scans/{id} ตรงที่ผูกด้วย session_id + qr_key แทน scan_id
# เพราะฝั่งแอพสแกน (index.html) เก็บ log ในเครื่องด้วย qr_key ไม่รู้จัก scan_id ของ DB
# อนุญาตให้สาขาลบได้เฉพาะ session ของสาขาตัวเอง และต้องยังไม่ปิดงาน (status != 'done')
# ถ้าปิดงานแล้วต้องให้ HQ ลบผ่าน /hq/scans/{id} แทน เพื่อให้มี oversight
# ════════════════════════════════════════════════════════════════
@app.delete("/scans")
def delete_scan_by_qr(session_id: int, qr_key: str, db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("SELECT id, branch_id, status FROM audit_sessions WHERE id = %s", (session_id,))
    sess = cur.fetchone()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    if user["role"] != "hq_admin" and sess["branch_id"] != user.get("branch_id"):
        raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์ลบ scan ของสาขาอื่น")
    if sess["status"] == "done" and user["role"] != "hq_admin":
        raise HTTPException(status_code=400, detail={
            "error": "session_closed",
            "message": "Session นี้ปิดงานแล้ว ลบจากหน้าแอพสแกนไม่ได้ กรุณาแจ้ง HQ ให้ลบให้"
        })
    cur.execute("""
        SELECT sl.id, sl.session_id, a.id AS asset_id, a.asset_code, a.name, a.qr_key
        FROM scan_logs sl
        JOIN assets a ON a.id = sl.asset_id
        WHERE sl.session_id = %s AND a.qr_key = %s
    """, (session_id, qr_key))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="ไม่พบ scan ของ QR นี้ในรอบนี้ (อาจยังไม่ถูกบันทึกเข้าระบบ)")
    _record_scan_delete(cur, dict(row), session_id, sess["branch_id"], user, source="branch")
    cur.execute("DELETE FROM scan_logs WHERE id = %s", (row["id"],))
    db.commit()
    return {"ok": True, "deleted_scan_id": row["id"]}

# ════════════════════════════════════════════════════════════════
# BRANCH SELF-SERVICE SCAN UPDATE (แก้บั๊ก "ถ่ายรูปแล้วอัพขึ้น Cloudinary สำเร็จ
# แต่ Dashboard ไม่มีรูป")
# สาเหตุจริง: index.html สร้าง scan_log ทันทีตอนสแกนเจอ QR (POST /scans ครั้งแรก,
# ยังไม่มีรูป) แล้วพอผู้ใช้ถ่ายรูป+กด SAVE จะยิง POST /scans อีกครั้งพร้อม photo_url
# ซึ่งชนกับ scan_log ที่มีอยู่แล้ว (asset เดิม, session เดิม) โดน guard "already_scanned"
# ตอบ 409 กลับไป แล้ว photo_url ก็หายไปเงียบๆ เพราะ frontend ปฏิบัติกับ 409 เป็น "สำเร็จ"
# (ดู reliableSend ใน index.html) ทางแก้: แยก endpoint นี้ไว้ "แก้ไข" record ที่มีอยู่แล้ว
# แทนที่จะพยายาม POST ซ้ำ — index.html เรียก endpoint นี้แทน saveScanLog() ตอนแนบรูป/remark
# ════════════════════════════════════════════════════════════════
@app.patch("/scans")
def update_scan_by_qr(session_id: int, qr_key: str, req: ScanLogPatchByQR,
                       db=Depends(get_db), user=Depends(get_current_user)):
    cur = db.cursor()
    cur.execute("SELECT id, branch_id, status FROM audit_sessions WHERE id = %s", (session_id,))
    sess = cur.fetchone()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    if user["role"] != "hq_admin" and sess["branch_id"] != user.get("branch_id"):
        raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์แก้ไข scan ของสาขาอื่น")
    if sess["status"] == "done" and user["role"] != "hq_admin":
        raise HTTPException(status_code=400, detail={
            "error": "session_closed",
            "message": "Session นี้ปิดงานแล้ว แก้ไขจากหน้าแอพสแกนไม่ได้ กรุณาแจ้ง HQ"
        })
    cur.execute("""
        SELECT sl.id FROM scan_logs sl
        JOIN assets a ON a.id = sl.asset_id
        WHERE sl.session_id = %s AND a.qr_key = %s
    """, (session_id, qr_key))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="ไม่พบ scan ของ QR นี้ในรอบนี้ (อาจยังไม่ถูกบันทึกเข้าระบบ)")

    fields, values = [], []
    if req.photo_url is not None: fields.append("photo_url = %s"); values.append(req.photo_url)
    if req.remark is not None: fields.append("remark = %s"); values.append(req.remark)
    if req.serial_found is not None: fields.append("serial_found = %s"); values.append(req.serial_found)
    if req.serial_match is not None: fields.append("serial_match = %s"); values.append(req.serial_match)
    if req.condition is not None: fields.append("condition = %s"); values.append(req.condition)
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    fields.append("updated_at = NOW()")
    values.append(row["id"])
    cur.execute(f"UPDATE scan_logs SET {', '.join(fields)} WHERE id = %s", values)
    db.commit()
    return {"ok": True, "scan_id": row["id"]}

@app.get("/hq/scan-delete-logs")
def hq_get_scan_delete_logs(branch_id: Optional[str] = None, session_id: Optional[int] = None,
                             db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    query = "SELECT * FROM scan_delete_logs WHERE 1=1"
    params = []
    if branch_id:
        query += " AND branch_id = %s"; params.append(branch_id)
    if session_id:
        query += " AND session_id = %s"; params.append(session_id)
    query += " ORDER BY deleted_at DESC LIMIT 200"
    cur.execute(query, tuple(params))
    return {"logs": [dict(r) for r in cur.fetchall()]}

@app.delete("/hq/scans/{scan_id}/photo", status_code=204)
async def hq_delete_scan_photo(scan_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("SELECT photo_url FROM scan_logs WHERE id = %s", (scan_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Scan not found")
    old_url = row["photo_url"]
    if old_url and "cloudinary.com" in old_url:
        try:
            cloud  = os.getenv("CLOUDINARY_CLOUD_NAME", "")
            key    = os.getenv("CLOUDINARY_API_KEY", "")
            secret = os.getenv("CLOUDINARY_API_SECRET", "")
            if cloud and key and secret:
                public_id = _cloudinary_public_id(old_url)
                if public_id:
                    ts  = int(time.time())
                    sig = _cloudinary_sign({"public_id": public_id, "timestamp": ts}, secret)
                    async with httpx.AsyncClient() as client:
                        await client.post(
                            f"https://api.cloudinary.com/v1_1/{cloud}/image/destroy",
                            data={"public_id": public_id, "timestamp": ts,
                                  "api_key": key, "signature": sig}, timeout=15)
        except Exception as e:
            print(f"[WARN] Cloudinary delete failed: {e}")
    cur.execute("UPDATE scan_logs SET photo_url = NULL, updated_at = NOW() WHERE id = %s", (scan_id,))
    db.commit()
    cur.close()

@app.post("/hq/scans/{scan_id}/photo")
async def hq_upload_scan_photo(scan_id: int, photo: UploadFile = File(...), db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    if not photo.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์รูปภาพ")
    content = await photo.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="ไฟล์ใหญ่เกิน 5MB")
    cur = db.cursor()
    cur.execute("""
        SELECT b.id as branch_id FROM scan_logs sl
        LEFT JOIN audit_sessions s ON s.id = sl.session_id
        LEFT JOIN branches b ON b.id = s.branch_id
        WHERE sl.id = %s
    """, (scan_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Scan not found")
    branch_id = row["branch_id"] or "unknown"
    cloud  = os.getenv("CLOUDINARY_CLOUD_NAME", "")
    key    = os.getenv("CLOUDINARY_API_KEY", "")
    secret = os.getenv("CLOUDINARY_API_SECRET", "")
    if not (cloud and key and secret):
        raise HTTPException(status_code=500, detail="Cloudinary ยังไม่ได้ตั้งค่า ENV")
    ts        = int(time.time())
    public_id = f"bata-audit/{branch_id}/scan_{scan_id}_{ts}"
    sig       = _cloudinary_sign({"public_id": public_id, "timestamp": ts}, secret)
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"https://api.cloudinary.com/v1_1/{cloud}/image/upload",
            data={"api_key": key, "timestamp": ts, "signature": sig, "public_id": public_id},
            files={"file": (photo.filename, content, photo.content_type)}, timeout=30)
    if resp.status_code != 200:
        raise HTTPException(status_code=500, detail=f"Upload ไม่สำเร็จ: {resp.text[:200]}")
    photo_url = resp.json().get("secure_url", "")
    if not photo_url:
        raise HTTPException(status_code=500, detail="ไม่ได้รับ URL จาก Cloudinary")
    cur.execute("UPDATE scan_logs SET photo_url = %s, updated_at = NOW() WHERE id = %s", (photo_url, scan_id))
    db.commit()
    cur.close()
    return {"photo_url": photo_url}

@app.patch("/hq/unmatched/{unmatched_id}")
def hq_update_unmatched(unmatched_id: int, req: UnmatchedUpdate, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    if req.status not in ("matched", "rejected"):
        raise HTTPException(status_code=400, detail="status must be 'matched' or 'rejected'")
    cur = db.cursor()
    cur.execute("""
        SELECT ua.*, u.full_name as auditor_name FROM unmatched_assets ua
        LEFT JOIN users u ON u.id = ua.scanned_by
        WHERE ua.id = %s
    """, (unmatched_id,))
    um = cur.fetchone()
    if not um:
        raise HTTPException(status_code=404, detail="Unmatched asset not found")
    hq_note = req.hq_note or f"{'Matched' if req.status == 'matched' else 'Rejected'} by HQ - {user['employee_id']}"
    cur.execute("""
        UPDATE unmatched_assets
        SET status=%s, hq_note=%s, matched_asset_code=%s,
            serial_no=COALESCE(%s,serial_no), reviewed_by=%s,
            reviewed_at=NOW(), updated_at=NOW()
        WHERE id=%s
    """, (req.status, hq_note, req.matched_asset_code, req.serial_no, user["user_id"], unmatched_id))
    if req.status == "matched" and req.matched_asset_code:
        cur.execute("SELECT id FROM assets WHERE asset_code = %s LIMIT 1", (req.matched_asset_code,))
        asset = cur.fetchone()
        if asset:
            cur.execute("""
                INSERT INTO scan_logs
                    (session_id, asset_id, scanned_by, serial_found, serial_match,
                     condition, remark, photo_url, hq_note, scanned_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT DO NOTHING
            """, (um["session_id"], asset["id"], um["scanned_by"],
                  req.serial_no or um.get("serial_no"), None, "good",
                  f"[Matched from Unmatched] QR: {um['scanned_qr']}",
                  um.get("photo_url"), hq_note,
                  um.get("scanned_at") or datetime.now()))
    db.commit()
    return {"ok": True, "unmatched_id": unmatched_id, "status": req.status}

@app.get("/hq/unmatched")
def hq_get_unmatched(db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("""
        SELECT ua.id, ua.scanned_qr, ua.name_guess, ua.serial_no, ua.photo_url,
               ua.scanned_at, ua.status, ua.hq_note, ua.matched_asset_code,
               u.email AS auditor, b.id AS branch_id, b.name AS branch_name
        FROM unmatched_assets ua
        LEFT JOIN users u ON u.id = ua.scanned_by
        LEFT JOIN audit_sessions s ON s.id = ua.session_id
        LEFT JOIN branches b ON b.id = s.branch_id
        ORDER BY ua.scanned_at DESC
    """)
    return {"items": [dict(r) for r in cur.fetchall()]}

@app.get("/hq/assets")
def hq_get_all_assets(db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("""
        SELECT a.id, a.qr_key, a.asset_code, a.seq, a.name, a.serial_no,
               a.location_code, a.purchase_date, a.status, a.qty,
               CASE WHEN sl.id IS NOT NULL THEN true ELSE false END AS is_scanned,
               sl.scanned_at, u.email AS scanned_by, s.branch_id
        FROM assets a
        LEFT JOIN scan_logs sl ON sl.asset_id = a.id
        LEFT JOIN users u ON u.id = sl.scanned_by
        LEFT JOIN audit_sessions s ON s.id = sl.session_id
        WHERE a.status = 'active' AND COALESCE(a.is_active, true) = true
        ORDER BY a.location_code, a.asset_code, a.seq
    """)
    assets = [dict(r) for r in cur.fetchall()]
    total   = len(assets)
    scanned = sum(1 for a in assets if a["is_scanned"])
    return {"total": total, "scanned": scanned, "pending": total - scanned, "assets": assets}

@app.delete("/hq/sessions/{session_id}")
def hq_delete_session(session_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("DELETE FROM scan_logs WHERE session_id = %s", (session_id,))
    cur.execute("DELETE FROM unmatched_assets WHERE session_id = %s", (session_id,))
    cur.execute("DELETE FROM audit_sessions WHERE id = %s RETURNING id", (session_id,))
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail="Session not found")
    db.commit()
    return {"ok": True, "session_id": session_id}

@app.post("/hq/sessions/merge")
def hq_merge_sessions(req: MergeSessionsRequest, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    if not req.merge_session_ids:
        raise HTTPException(status_code=400, detail="ไม่มี session ที่ต้อง merge")
    cur = db.cursor()
    cur.execute("SELECT id FROM audit_sessions WHERE id = %s", (req.primary_session_id,))
    if not cur.fetchone():
        raise HTTPException(status_code=404, detail=f"Primary session #{req.primary_session_id} ไม่พบ")
    moved_scans = moved_unmatched = deleted_sessions = 0
    for sid in req.merge_session_ids:
        if sid == req.primary_session_id:
            continue
        cur.execute("""
            UPDATE scan_logs SET session_id=%s
            WHERE session_id=%s
              AND asset_id NOT IN (SELECT asset_id FROM scan_logs WHERE session_id=%s)
        """, (req.primary_session_id, sid, req.primary_session_id))
        moved_scans += cur.rowcount
        cur.execute("DELETE FROM scan_logs WHERE session_id = %s", (sid,))
        cur.execute("UPDATE unmatched_assets SET session_id=%s WHERE session_id=%s", (req.primary_session_id, sid))
        moved_unmatched += cur.rowcount
        cur.execute("DELETE FROM audit_sessions WHERE id = %s", (sid,))
        deleted_sessions += cur.rowcount
    cur.execute("""
        UPDATE audit_sessions SET status='on_process'
        WHERE id=%s AND status='open'
          AND EXISTS (SELECT 1 FROM scan_logs WHERE session_id=%s)
    """, (req.primary_session_id, req.primary_session_id))
    db.commit()
    cur.close()
    return {"ok": True, "primary_session_id": req.primary_session_id,
            "deleted_sessions": deleted_sessions, "moved_scans": moved_scans,
            "moved_unmatched": moved_unmatched}

# ════════════════════════════════════════════════════════════════
# ASSET IMPORT (HQ only)
# ════════════════════════════════════════════════════════════════
EXPECTED_COLUMNS = [
    "รหัสสินทรัพย์", "ลำดับ", "รหัสที่ตั้ง", "ชื่อที่ตั้ง",
    "รายละเอียดสินทรัพย์(อังกฤษ)", "สถานะสินทรัพย์", "เลขที่เครื่อง",
    "วันที่ซื้อ", "จำนวน", "มูลค่าสินทรัพย์", "คสส.", "มูลค่าทางบัญชี",
]

# รหัสสถานะในไฟล์ Excel ของ HQ ('A' = active เป็นต้น) -> ค่า status เต็มที่ระบบใช้กรองอยู่จริง (WHERE status='active')
# ถ้าเจอรหัสใหม่ที่ไม่อยู่ใน map นี้ ให้เพิ่มเข้ามาแทนที่จะปล่อยรหัสดิบเข้า DB ตรงๆ (จะทำให้ asset หายจากทุกหน้าที่กรอง status='active')
_STATUS_CODE_MAP = {
    "A": "active",
    "D": "disposed",
    "T": "transferred",
    "M": "missing",
}

def _parse_thai_date_string(value):
    """แปลง string วันที่ที่มาจากไฟล์ Excel (เก็บเป็น text ไม่ใช่ date serial) ให้เป็น ISO YYYY-MM-DD
    รองรับรูปแบบที่เจอจริง: 26/06/2026 (DD/MM/YYYY ค.ศ.), 26/06/2569 (DD/MM/YYYY พ.ศ.)
    คืนค่า (parsed_date_or_None, ok:bool) — ok=False หมายถึง parse ไม่ได้ ต้องเตือนแอดมิน ไม่ใช่ปล่อยเป็น NULL เงียบๆ"""
    if not value:
        return None, True  # ไม่มีค่าเลยถือว่าโอเค ไม่ใช่ error การ parse
    s = str(value).strip()
    if not s:
        return None, True
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s, fmt)
            year = dt.year
            if year > 2400:  # พ.ศ. -> ค.ศ.
                year -= 543
            return f"{year:04d}-{dt.month:02d}-{dt.day:02d}", True
        except ValueError:
            continue
    # รูปแบบไม่ตรง pattern ใดเลย — ส่งกลับ None พร้อม ok=False เพื่อให้ฝั่ง preview แสดง warning แทนที่จะเงียบ
    return None, False


def _parse_asset_excel(content: bytes, filename: str):
    """อ่านไฟล์ .xls/.xlsx แล้วคืน list of dict ที่ map เป็นชื่อ column ของ DB แล้ว"""
    rows_out = []

    def _row_to_dict(values):
        if not values or not str(values[0]).strip():
            return None
        if str(values[0]).strip().lower() in ("grand total", "รหัสสินทรัพย์"):
            return None
        try:
            asset_code = str(values[0]).strip()
            seq        = str(values[1]).strip() if values[1] not in (None, "") else ""
            location_code = str(values[2]).strip() if len(values) > 2 else ""
            location_name = str(values[3]).strip() if len(values) > 3 else ""
            name          = str(values[4]).strip() if len(values) > 4 else ""
            status_code   = str(values[5]).strip() if len(values) > 5 else "A"
            status        = _STATUS_CODE_MAP.get(status_code.upper(), "active")
            status_known  = status_code.upper() in _STATUS_CODE_MAP
            serial_no     = str(values[6]).strip() if len(values) > 6 and values[6] else None
            purchase_date_raw = values[7] if len(values) > 7 else None
            qty            = float(values[8]) if len(values) > 8 and values[8] not in (None, "") else 0
            purchase_price = float(values[9]) if len(values) > 9 and values[9] not in (None, "") else 0
            accumulated_dep = float(values[10]) if len(values) > 10 and values[10] not in (None, "") else 0
            net_book_value  = float(values[11]) if len(values) > 11 and values[11] not in (None, "") else 0
        except (ValueError, IndexError) as e:
            raise HTTPException(status_code=400, detail=f"แถวข้อมูลผิดรูปแบบ: {values} ({e})")
        if not asset_code:
            return None
        warnings = []
        if not status_known:
            warnings.append(f"ไม่รู้จักรหัสสถานะ '{status_code}' — ใช้ 'active' เป็นค่า default ไว้ก่อน กรุณาตรวจสอบ")
        return {
            "asset_code": asset_code, "seq": seq, "qr_key": asset_code + seq,
            "location_code": location_code, "location_name": location_name,
            "name": name, "status": status, "serial_no": serial_no,
            "purchase_date_raw": purchase_date_raw,
            "qty": qty, "purchase_price": purchase_price,
            "accumulated_dep": accumulated_dep, "net_book_value": net_book_value,
            "warnings": warnings,
        }

    if filename.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=content)
        sheet = wb.sheet_by_index(0)
        for r in range(1, sheet.nrows):
            values = sheet.row_values(r)
            d = _row_to_dict(values)
            if d:
                if isinstance(d["purchase_date_raw"], float):
                    y, m, day, *_ = xlrd.xldate_as_tuple(d["purchase_date_raw"], wb.datemode)
                    d["purchase_date"] = f"{y:04d}-{m:02d}-{day:02d}"
                else:
                    parsed, ok = _parse_thai_date_string(d["purchase_date_raw"])
                    d["purchase_date"] = parsed
                    if not ok:
                        d["warnings"].append(f"แปลงวันที่ '{d['purchase_date_raw']}' ไม่ได้ — purchase_date จะถูกบันทึกเป็นค่าว่าง")
                rows_out.append(d)
    elif filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.worksheets[0]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            d = _row_to_dict(list(row))
            if d:
                pd = d["purchase_date_raw"]
                if hasattr(pd, "strftime"):
                    d["purchase_date"] = pd.strftime("%Y-%m-%d")
                else:
                    parsed, ok = _parse_thai_date_string(pd)
                    d["purchase_date"] = parsed
                    if not ok:
                        d["warnings"].append(f"แปลงวันที่ '{pd}' ไม่ได้ — purchase_date จะถูกบันทึกเป็นค่าว่าง")
                rows_out.append(d)
    else:
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์ .xls หรือ .xlsx เท่านั้น")

    if not rows_out:
        raise HTTPException(status_code=400, detail="ไม่พบข้อมูลในไฟล์ (หรือไฟล์ว่างเปล่า)")
    return rows_out


IMPORT_FILE_SIZE_LIMIT = 10 * 1024 * 1024  # 10MB — asset list ไฟล์จริงมักไม่ใหญ่เกินนี้ ป้องกันอัปโหลดไฟล์ผิดประเภท

async def _read_upload_with_size_limit(file: UploadFile) -> bytes:
    content = await file.read()
    if len(content) > IMPORT_FILE_SIZE_LIMIT:
        raise HTTPException(status_code=400, detail=f"ไฟล์ใหญ่เกินไป ({len(content)/1024/1024:.1f}MB) — รองรับสูงสุด {IMPORT_FILE_SIZE_LIMIT//1024//1024}MB")
    return content


@app.post("/hq/assets/import/preview")
async def hq_import_assets_preview(file: UploadFile = File(...), user=Depends(get_current_user), db=Depends(get_db)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    content = await _read_upload_with_size_limit(file)
    rows = _parse_asset_excel(content, file.filename)

    cur = db.cursor()
    codes = [(r["asset_code"], r["seq"]) for r in rows]
    if codes:
        # FIX: เดิมเช็ค duplicate จากทุกแถวไม่ว่า is_active หรือไม่ ทำให้ asset ที่เคยถูก
        # Undo import ไปแล้ว (is_active=false) ยังถูกฟ้องว่า "ซ้ำ" อยู่ตลอดไป ทั้งที่จริงๆ
        # ควร import ใหม่ได้ (ดูคอมเมนต์ยาวใน /hq/assets/import/confirm ด้านล่าง)
        cur.execute("SELECT asset_code, seq FROM assets WHERE (asset_code, seq) IN %s AND COALESCE(is_active, true) = true", (tuple(codes),))
        existing = {(r["asset_code"], r["seq"]) for r in cur.fetchall()}
    else:
        existing = set()

    preview = []
    for r in rows:
        is_dup = (r["asset_code"], r["seq"]) in existing
        preview.append({**{k: v for k, v in r.items() if k != "purchase_date_raw"}, "duplicate": is_dup})

    return {"file_name": file.filename, "total_rows": len(rows),
            "duplicate_count": sum(1 for p in preview if p["duplicate"]),
            "warning_count": sum(1 for p in preview if p["warnings"]),
            "rows": preview}


@app.post("/hq/assets/import/confirm")
async def hq_import_assets_confirm(file: UploadFile = File(...), user=Depends(get_current_user), db=Depends(get_db)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    content = await _read_upload_with_size_limit(file)
    rows = _parse_asset_excel(content, file.filename)

    batch_id = uuid.uuid4().hex
    cur = db.cursor()
    inserted = skipped = 0
    try:
        for r in rows:
            # FIX (root cause of "import สำเร็จแต่หาไม่เจอ" ที่เกิดซ้ำๆ): เดิมใช้
            # ON CONFLICT (asset_code, seq) DO NOTHING เฉยๆ — ปัญหาคือ "Undo import"
            # (ดู /hq/assets/import-batches/{batch_id}/undo ด้านล่าง) แค่ set is_active=false
            # ไม่เคยลบแถวจริง ดังนั้นถ้าเคย import แล้ว undo ไป asset_code+seq นั้นจะชน
            # unique constraint นี้ตลอดไป โดน DO NOTHING skip ทุกครั้งที่ import ซ้ำ
            # (ต่อให้เป็นไฟล์ถูกต้อง คนละไฟล์ก็ตาม) และเพราะ is_active=false มันจะไม่โผล่ใน
            # /hq/assets เลย กลายเป็น "หายไปถาวร" ทั้งที่ user ไม่ได้ทำอะไรผิด
            # ทางแก้: ถ้า conflict เกิดกับแถวที่ is_active=false อยู่แล้ว (เคยถูก undo)
            # ให้ reactivate + อัปเดตข้อมูลใหม่แทนที่จะ skip เงียบๆ — ส่วนแถวที่ยัง
            # is_active=true (ซ้ำจริงกับของที่ใช้งานอยู่) ยังคง skip เหมือนเดิม
            cur.execute("""
                INSERT INTO assets
                    (asset_code, seq, qr_key, location_code, location_name, name,
                     status, serial_no, purchase_date, qty, purchase_price,
                     accumulated_dep, net_book_value, import_batch_id, is_active, imported_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,true,NOW())
                ON CONFLICT (asset_code, seq) DO UPDATE SET
                    qr_key = EXCLUDED.qr_key,
                    location_code = EXCLUDED.location_code,
                    location_name = EXCLUDED.location_name,
                    name = EXCLUDED.name,
                    status = EXCLUDED.status,
                    serial_no = EXCLUDED.serial_no,
                    purchase_date = EXCLUDED.purchase_date,
                    qty = EXCLUDED.qty,
                    purchase_price = EXCLUDED.purchase_price,
                    accumulated_dep = EXCLUDED.accumulated_dep,
                    net_book_value = EXCLUDED.net_book_value,
                    import_batch_id = EXCLUDED.import_batch_id,
                    is_active = true,
                    imported_at = NOW()
                WHERE assets.is_active = false
            """, (r["asset_code"], r["seq"], r["qr_key"], r["location_code"], r["location_name"],
                  r["name"], r["status"], r["serial_no"], r["purchase_date"], r["qty"],
                  r["purchase_price"], r["accumulated_dep"], r["net_book_value"], batch_id))
            if cur.rowcount:
                inserted += 1
            else:
                skipped += 1

        cur.execute("""
            INSERT INTO import_batches (batch_id, file_name, row_count, imported_by, status)
            VALUES (%s,%s,%s,%s,'active')
        """, (batch_id, file.filename, inserted, user["employee_id"]))
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import ล้มเหลว ไม่มีข้อมูลเข้าระบบ: {e}")

    return {"ok": True, "batch_id": batch_id, "inserted": inserted, "skipped_duplicates": skipped}


@app.get("/hq/assets/import-batches")
def hq_list_import_batches(db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("SELECT * FROM import_batches ORDER BY imported_at DESC LIMIT 50")
    return {"batches": [dict(r) for r in cur.fetchall()]}


@app.post("/hq/assets/import-batches/{batch_id}/undo")
def hq_undo_import_batch(batch_id: str, force: bool = False, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    cur = db.cursor()
    cur.execute("SELECT status FROM import_batches WHERE batch_id = %s", (batch_id,))
    batch = cur.fetchone()
    if not batch:
        raise HTTPException(status_code=404, detail="ไม่พบ import batch นี้")
    if batch["status"] == "undone":
        raise HTTPException(status_code=400, detail="Batch นี้ถูก undo ไปแล้ว")

    # เช็คก่อนว่ามี scan_logs ผูกกับ asset ใน batch นี้แล้วหรือยัง — ถ้ามีและไม่ force
    # ให้แค่เตือนกลับไป ไม่ undo ทันที เพราะถ้า undo แล้วจะเหลือ scan_logs.asset_id
    # ที่ชี้ไปยัง asset ที่ is_active=false (ค้างเป็น dangling reference)
    cur.execute("""
        SELECT COUNT(*) AS cnt FROM scan_logs sl
        JOIN assets a ON a.id = sl.asset_id
        WHERE a.import_batch_id = %s
    """, (batch_id,))
    scan_count = cur.fetchone()["cnt"]
    if scan_count > 0 and not force:
        raise HTTPException(status_code=409, detail={
            "error": "has_linked_scans",
            "message": f"พบ scan log ที่ผูกกับ asset ใน batch นี้แล้ว {scan_count} รายการ — ถ้า undo ต่อ scan log เหล่านี้จะยังอ้างอิงถึง asset ที่ถูกซ่อนไป กรุณายืนยันอีกครั้งถ้าต้องการ undo จริงๆ",
            "scan_count": scan_count,
        })

    cur.execute("UPDATE assets SET is_active = false WHERE import_batch_id = %s", (batch_id,))
    undone_count = cur.rowcount
    cur.execute("UPDATE import_batches SET status = 'undone' WHERE batch_id = %s", (batch_id,))
    db.commit()
    return {"ok": True, "batch_id": batch_id, "undone_rows": undone_count, "linked_scans_warned": scan_count}


# ════════════════════════════════════════════════════════════════
# BRANDED EXCEL EXPORT
# เดิม Export Excel ทำฝั่ง browser ล้วน (SheetJS ฟรี) ซึ่งใส่สี/ตัวหนา/โลโก้ไม่ได้จริง
# (เป็นข้อจำกัดของ library ฟรีตัวนั้น ไม่รองรับ write style กลับเข้าไฟล์ xlsx)
# ย้ายมาสร้างไฟล์ฝั่ง backend ด้วย openpyxl แทน เพราะใส่สี/ฟอนต์/border/freeze panes
# ได้เต็มรูปแบบ — frontend ส่งแค่ headers/rows ที่ filter ไว้แล้วมาให้ endpoint นี้จัดหน้าและคืนไฟล์
# ════════════════════════════════════════════════════════════════
BATA_RED = "C8102E"
BATA_RED_DARK = "AB0D27"
LIGHT_GRAY = "F5F5F7"

def _build_branded_workbook(req: ExportExcelRequest, exported_by: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    n_cols = max(len(req.headers), 1)
    last_col_letter = get_column_letter(n_cols)

    thin = Side(style="thin", color="D0D0D5")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # แถว 1-2: กล่องโลโก้ "B" สีแดง (เลียนแบบ brand mark ที่ใช้ใน Dashboard จริง — ไม่มีไฟล์รูปโลโก้แยก
    # จึงสร้างเป็นเซลล์สีแดงตัวอักษร B แทน ให้ผลลัพธ์เหมือนกันโดยไม่ต้อง embed รูปภาพ)
    ws.merge_cells("A1:A2")
    logo_cell = ws["A1"]
    logo_cell.value = "B"
    logo_cell.font = Font(name="Arial", size=24, bold=True, color="FFFFFF")
    logo_cell.fill = PatternFill("solid", fgColor=BATA_RED)
    logo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # ชื่อบริษัทเต็ม + ชื่อรายงาน ต่อจากกล่องโลโก้
    ws.merge_cells(f"B1:{last_col_letter}1")
    title_cell = ws["B1"]
    title_cell.value = "Bata (Thailand) Co., Ltd. — Fixed Asset Audit"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=BATA_RED_DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"B2:{last_col_letter}2")
    subtitle_cell = ws["B2"]
    subtitle_cell.value = req.title
    subtitle_cell.font = Font(name="Arial", size=11, bold=True, color="333333")
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    # แถวข้อมูล meta (สาขา / วันที่ export / จำนวน / export โดยใคร)
    meta_rows = [
        ("Branch", req.branch_label),
        ("Exported At", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Exported By", exported_by),
        ("Total Items", str(len(req.rows))),
    ]
    r = 3
    for label, value in meta_rows:
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=9, bold=True, color="777777")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
        ws.cell(row=r, column=2, value=value).font = Font(name="Arial", size=9, color="333333")
        r += 1
    r += 1  # แถวว่างคั่น

    header_row = r
    for c, h in enumerate(req.headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BATA_RED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    r += 1

    for i, row_data in enumerate(req.rows):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border_all
            cell.font = Font(name="Arial", size=10, color="222222")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        r += 1

    r += 2  # แถวว่างก่อนช่องเซ็นชื่อ
    ws.cell(row=r, column=1, value="ผู้ตรวจนับ (Auditor)").font = Font(name="Arial", size=9, bold=True)
    sig_col_2 = max(3, (n_cols // 2) + 1)
    ws.cell(row=r, column=sig_col_2, value="ผู้ตรวจสอบ / รับรอง (HQ)").font = Font(name="Arial", size=9, bold=True)
    r += 3
    ws.cell(row=r, column=1, value="ลงชื่อ ..................................... วันที่ ..................").font = Font(name="Arial", size=9)
    ws.cell(row=r, column=sig_col_2, value="ลงชื่อ ..................................... วันที่ ..................").font = Font(name="Arial", size=9)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.column_dimensions["A"].width = 20
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.post("/hq/export/excel")
def hq_export_excel(req: ExportExcelRequest, user=Depends(get_current_user)):
    if user["role"] != "hq_admin":
        raise HTTPException(status_code=403, detail="HQ admin only")
    if not req.headers or req.rows is None:
        raise HTTPException(status_code=400, detail="ต้องมี headers และ rows")
    buf = _build_branded_workbook(req, exported_by=user.get("employee_id") or user.get("full_name") or "—")
    filename = req.filename or "export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ════════════════════════════════════════════════════════════════
# HEALTH CHECK
# ════════════════════════════════════════════════════════════════
@app.get("/")
def root():
    return {"status": "ok", "service": "Bata Asset Audit API", "version": "1.2.0"}

@app.get("/health")
def health(db=Depends(get_db)):
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) as cnt FROM assets")
    cnt = cur.fetchone()["cnt"]
    return {"status": "ok", "assets_in_db": cnt}
